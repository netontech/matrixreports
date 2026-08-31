"""Excel output: durations stay summable, layout scales with the data."""

from datetime import date, timedelta

import openpyxl
import pytest

from matrixreports.builder import AttendanceBuilder
from matrixreports.duration import hhmm
from matrixreports.excel import write_csv, write_workbook
from matrixreports.reports import build_daily_report, build_summary_report

from conftest import DAY


@pytest.fixture
def book(config, source):
    return AttendanceBuilder(config, source).build(DAY, DAY)


def test_workbook_has_a_column_per_break(book, tmp_path):
    table = build_daily_report(book, DAY)
    path = write_workbook(table, tmp_path / "daily.xlsx")
    worksheet = openpyxl.load_workbook(path)["daily"]

    header_row = next(
        row for row in worksheet.iter_rows(values_only=True)
        if row and row[0] == "#"
    )
    assert list(header_row).count("MINS") == 11
    assert list(header_row).count("OUT") == 11


def test_durations_are_written_as_summable_excel_times(book, tmp_path):
    table = build_daily_report(book, DAY)
    path = write_workbook(table, tmp_path / "daily.xlsx")
    worksheet = openpyxl.load_workbook(path)["daily"]

    headers = [cell.value for cell in worksheet[_header_row_index(worksheet)]]
    column = headers.index("Actual Works Hours") + 1
    cell = worksheet.cell(row=_header_row_index(worksheet) + 1, column=column)

    # Stored numerically with a duration format, so Excel can SUM the column;
    # openpyxl reads such a cell back as a timedelta, not as the string "8.13"
    # and not as the decimal 8.13 the legacy sheets use.
    assert cell.number_format == "[h]:mm"
    assert isinstance(cell.value, timedelta)
    assert hhmm(cell.value) == "08:13"
    assert _raw_cell_is_numeric(path, "daily", cell.coordinate)


def test_two_duration_cells_add_up_correctly_in_the_sheet(config, employees, tmp_path):
    """The property the legacy hh.mm decimals lose."""
    from matrixreports.datasource import InMemoryPunchSource
    from matrixreports.reports import build_weekly_report
    from conftest import punches

    data = {
        "2": punches("2", ["10:00", "18:49"], date(2026, 6, 1))
        + punches("2", ["10:00", "18:45"], date(2026, 6, 2))
    }
    source = InMemoryPunchSource(employees[:2], data)
    book = AttendanceBuilder(config, source).build(date(2026, 6, 1), date(2026, 6, 2))
    path = write_workbook(build_weekly_report(book), tmp_path / "weekly.xlsx")

    worksheet = openpyxl.load_workbook(path)["weekly"]
    header_index = _header_row_index(worksheet)
    headers = [cell.value for cell in worksheet[header_index]]
    total_column = headers.index("Total Hours") + 1
    row = next(
        index for index in range(header_index + 1, worksheet.max_row + 1)
        if worksheet.cell(row=index, column=2).value == "Light Walker"
    )
    stored = worksheet.cell(row=row, column=total_column).value
    assert hhmm(stored) == "17:34"
    # 8:49 + 8:45. A decimal sheet would show 17.94 here.
    assert stored == timedelta(hours=17, minutes=34)


def test_status_codes_are_written_as_text(book, tmp_path):
    from matrixreports.reports import build_monthly_report

    path = write_workbook(build_monthly_report(book), tmp_path / "monthly.xlsx")
    worksheet = openpyxl.load_workbook(path)["monthly"]
    values = {
        cell.value
        for row in worksheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    }
    assert "Sick Leave" in values


def test_summary_sections_stack_on_one_sheet(book, tmp_path):
    sections = build_summary_report(book, DAY)
    path = write_workbook(sections, tmp_path / "summary.xlsx",
                          combine=True, combined_sheet_name="Summary")
    workbook = openpyxl.load_workbook(path)
    assert workbook.sheetnames == ["Summary"]
    text = {
        cell.value
        for row in workbook["Summary"].iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    }
    assert any("Late IN" in value for value in text)
    assert any("Not Present" in value for value in text)


def test_csv_export_renders_durations_as_hhmm(book, tmp_path):
    table = build_daily_report(book, DAY)
    path = write_csv(table, tmp_path / "daily.csv")
    lines = path.read_text(encoding="utf-8").splitlines()
    assert lines[0].count("MINS") == 11
    assert "08:13" in lines[1]


def test_writing_nothing_is_an_error(tmp_path):
    with pytest.raises(ValueError, match="no report tables"):
        write_workbook([], tmp_path / "empty.xlsx")


def _raw_cell_is_numeric(path, sheet_name: str, coordinate: str) -> bool:
    """Look inside the .xlsx to confirm the cell holds a number, not text.

    A cell written as text would carry t="s"/t="inlineStr"; a numeric cell has no
    type attribute. This is what makes the column summable inside Excel.
    """
    import re
    import zipfile

    with zipfile.ZipFile(path) as archive:
        workbook = archive.read("xl/workbook.xml").decode()
        order = re.findall(r'<sheet[^>]*name="([^"]+)"', workbook)
        index = order.index(sheet_name) + 1
        xml = archive.read(f"xl/worksheets/sheet{index}.xml").decode()
    match = re.search(rf'<c r="{coordinate}"([^>]*)>(.*?)</c>', xml)
    assert match, f"cell {coordinate} not found in sheet XML"
    attributes, body = match.groups()
    return 't="s"' not in attributes and "<v>" in body


def _header_row_index(worksheet) -> int:
    for row in worksheet.iter_rows():
        if row[0].value == "#":
            return row[0].row
    raise AssertionError("header row not found")
