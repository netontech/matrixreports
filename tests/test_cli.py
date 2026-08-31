"""CLI: the commands run end to end against the SQLite fixture."""

from __future__ import annotations

import yaml

import openpyxl
import pytest

from matrixreports.cli import main

from conftest import DAY


@pytest.fixture
def config_file(tmp_path, sqlite_db):
    path = tmp_path / "config.yaml"
    path.write_text(
        yaml.safe_dump(
            {
                "company": {"name": "Test Co"},
                "database": {"driver": "sqlite", "path": str(sqlite_db)},
                "schema": {
                    "employees": {
                        "table": "EmployeeMaster",
                        "columns": {"id": "EmployeeID", "code": "EmployeeCode",
                                    "name": "EmployeeName", "department": "Department"},
                        "where": "IsActive = 1",
                    },
                    "punches": {
                        "table": "AttendanceLog",
                        "columns": {"emp_id": "EmployeeID", "timestamp": "PunchDateTime",
                                    "direction": "InOutFlag", "device": "DoorName"},
                    },
                    "leave": {
                        "table": "LeaveRegister",
                        "columns": {"emp_id": "EmployeeID", "date_from": "LeaveFrom",
                                    "date_to": "LeaveTo", "code": "LeaveType"},
                    },
                    "direction_in": ["1"],
                    "direction_out": ["2"],
                },
                "shift": {"start": "10:00", "end": "19:00", "weekly_off_days": [4]},
                "leave_codes": {"SL": "Sick Leave"},
            }
        ),
        encoding="utf-8",
    )
    return path


def run(config_file, *args) -> int:
    return main(["--config", str(config_file), *args])


def test_check_reports_punch_depth(config_file, capsys):
    assert run(config_file, "check", "--from", "2026-06-01") == 0
    output = capsys.readouterr().out
    assert "24 punches, 11 breaks" in output
    assert "dropped by the stock report" in output
    assert "11 OUT/IN groups" in output


def test_daily_command_writes_a_workbook(config_file, tmp_path, capsys):
    out = tmp_path / "daily.xlsx"
    assert run(config_file, "--out", str(out), "daily", "--date", "2026-06-01") == 0
    assert out.exists()
    worksheet = openpyxl.load_workbook(out)["daily"]
    headers = next(
        row for row in worksheet.iter_rows(values_only=True) if row and row[0] == "#"
    )
    assert list(headers).count("MINS") == 11
    assert "24 punches, 11 breaks" in capsys.readouterr().out


def test_groups_flag_pins_the_layout(config_file, tmp_path):
    out = tmp_path / "pinned.xlsx"
    assert run(config_file, "--out", str(out), "--groups", "6",
               "daily", "--date", "2026-06-01") == 0
    worksheet = openpyxl.load_workbook(out)["daily"]
    headers = next(
        row for row in worksheet.iter_rows(values_only=True) if row and row[0] == "#"
    )
    assert list(headers).count("MINS") == 6


def test_summary_command(config_file, tmp_path):
    out = tmp_path / "summary.xlsx"
    assert run(config_file, "--out", str(out), "summary", "--date", "2026-06-01") == 0
    assert openpyxl.load_workbook(out).sheetnames == ["Summary"]


def test_monthly_and_weekly_and_yearly_commands(config_file, tmp_path):
    for args, name in (
        (["monthly", "--month", "2026-06"], "monthly"),
        (["weekly", "--from", "2026-06-01", "--to", "2026-06-30"], "weekly"),
        (["yearly", "--year", "2026"], "yearly"),
    ):
        out = tmp_path / f"{name}.xlsx"
        assert run(config_file, "--out", str(out), *args) == 0
        assert name in openpyxl.load_workbook(out).sheetnames


def test_csv_output_format(config_file, tmp_path):
    out = tmp_path / "daily.csv"
    assert run(config_file, "--format", "csv", "--out", str(out),
               "daily", "--date", "2026-06-01") == 0
    assert out.exists()
    assert out.read_text(encoding="utf-8").splitlines()[0].count("MINS") == 11


def test_employee_filter(config_file, tmp_path):
    out = tmp_path / "one.xlsx"
    assert run(config_file, "--out", str(out), "--employee", "2",
               "daily", "--date", "2026-06-01") == 0
    worksheet = openpyxl.load_workbook(out)["daily"]
    header_row = next(
        row[0].row for row in worksheet.iter_rows() if row[0].value == "#"
    )
    names = [
        row[1]
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True)
        if row and row[1]
    ]
    assert names == ["Light Walker"]


def test_missing_config_exits_with_guidance(tmp_path, capsys):
    with pytest.raises(SystemExit) as excinfo:
        main(["--config", str(tmp_path / "nope.yaml"), "daily", "--date", "2026-06-01"])
    assert excinfo.value.code == 2
    assert "docs/schema-mapping.md" in capsys.readouterr().err


def test_bad_schema_mapping_reports_the_missing_column(tmp_path, sqlite_db, capsys):
    import yaml as _yaml

    path = tmp_path / "broken.yaml"
    path.write_text(
        _yaml.safe_dump(
            {
                "database": {"driver": "sqlite", "path": str(sqlite_db)},
                "schema": {
                    "employees": {
                        "table": "EmployeeMaster",
                        "columns": {"id": "EmployeeID", "name": "EmployeeName"},
                    },
                    # `timestamp` deliberately left unmapped
                    "punches": {"table": "AttendanceLog",
                                "columns": {"emp_id": "EmployeeID"}},
                },
            }
        ),
        encoding="utf-8",
    )
    assert main(["--config", str(path), "check", "--from", "2026-06-01"]) == 2
    assert "not mapped" in capsys.readouterr().err


def test_wrong_table_name_reports_a_helpful_error(tmp_path, sqlite_db, capsys):
    import yaml as _yaml

    path = tmp_path / "wrongtable.yaml"
    path.write_text(
        _yaml.safe_dump(
            {
                "database": {"driver": "sqlite", "path": str(sqlite_db)},
                "schema": {
                    "employees": {
                        "table": "NoSuchTable",
                        "columns": {"id": "EmployeeID", "name": "EmployeeName"},
                    },
                    "punches": {
                        "table": "AttendanceLog",
                        "columns": {"emp_id": "EmployeeID", "timestamp": "PunchDateTime"},
                    },
                },
            }
        ),
        encoding="utf-8",
    )
    assert main(["--config", str(path), "check", "--from", "2026-06-01"]) == 2
    assert "docs/schema-mapping.md" in capsys.readouterr().err


def test_discover_writes_a_working_config(config_file, tmp_path, capsys):
    out = tmp_path / "generated.yaml"
    assert run(config_file, "discover", "--write", str(out)) == 0
    assert out.exists()

    captured = capsys.readouterr()
    assert "Punch log" in captured.err
    assert "AttendanceLog" in captured.err

    # The generated config drives a real run with no hand editing.
    assert main(["--config", str(out), "check", "--from", "2026-06-01"]) == 0
    assert "24 punches, 11 breaks" in capsys.readouterr().out


def test_discover_refuses_to_overwrite(config_file, tmp_path):
    out = tmp_path / "existing.yaml"
    out.write_text("# do not clobber\n", encoding="utf-8")
    assert run(config_file, "discover", "--write", str(out)) == 2
    assert out.read_text(encoding="utf-8") == "# do not clobber\n"


def test_discover_prints_to_stdout_without_write(config_file, capsys):
    assert run(config_file, "discover") == 0
    assert "schema:" in capsys.readouterr().out


def test_discover_needs_a_database_not_a_csv(config_file, tmp_path, capsys):
    extract = tmp_path / "punches.csv"
    extract.write_text("emp_id,name,timestamp\n1,A,2026-06-01 10:00:00\n", encoding="utf-8")
    assert run(config_file, "--csv", str(extract), "discover") == 2
    assert "cannot inspect a CSV" in capsys.readouterr().err
