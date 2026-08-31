"""Write :class:`ReportTable` objects to .xlsx (and CSV).

Two habits from the legacy workbooks are deliberately not reproduced:

* durations are written as Excel *time* values with a ``[h]:mm`` format rather
  than as ``8.14``-style decimals, so a column of them can be summed in the
  sheet and still shows ``HH:MM`` past 24 hours;
* a column holds one type.  The supplied weekly sheet mixes ``'54:00'`` text
  with ``54.0`` numbers in the same column, which silently breaks both ``SUM``
  and sorting.
"""

from __future__ import annotations

import csv
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .duration import excel_time, hhmm
from .reports.base import Cell, ReportTable

NUMBER_FORMATS = {
    "time": "hh:mm",
    "duration": "[h]:mm",
    "date": "dd-mmm-yyyy",
    "int": "0",
    "number": "0.00",
    "percent": "0%",
    "text": "@",
    "code": "@",
}

TITLE_FONT = Font(bold=True, size=14)
COMPANY_FONT = Font(bold=True, size=12)
SUBTITLE_FONT = Font(italic=True, size=10)
HEADER_FONT = Font(bold=True, color="FFFFFF")
GROUP_FONT = Font(bold=True)
NOTE_FONT = Font(italic=True, size=9, color="9C5700")

HEADER_FILL = PatternFill("solid", fgColor="2F4F6F")
GROUP_FILL = PatternFill("solid", fgColor="DCE6F1")
CODE_FILLS = {
    "OFF": PatternFill("solid", fgColor="EDEDED"),
    "HOL": PatternFill("solid", fgColor="FFF2CC"),
    "A": PatternFill("solid", fgColor="FFC7CE"),
    "Absent": PatternFill("solid", fgColor="FFC7CE"),
}
LEAVE_FILL = PatternFill("solid", fgColor="E2EFDA")

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _excel_value(cell: Cell):
    """Convert a report cell into something openpyxl will store faithfully."""
    value = cell.value
    if value is None or value == "":
        return None
    if cell.style == "duration":
        return excel_time(value) if isinstance(value, timedelta) else value
    if cell.style == "time":
        if isinstance(value, datetime):
            # Stored as a full datetime so the date survives for night shifts;
            # the hh:mm format keeps the display identical to the old sheet.
            return value
        return value
    if cell.style == "date":
        return value
    if cell.style in {"int", "number", "percent"}:
        return value
    return str(value)


def _write_table(worksheet: Worksheet, table: ReportTable, *, start_row: int = 1) -> int:
    """Render one table onto a sheet. Returns the next free row."""
    row = start_row
    width = max(len(table.columns), 1)

    if table.company:
        worksheet.cell(row=row, column=1, value=table.company).font = COMPANY_FONT
        row += 1
    worksheet.cell(row=row, column=1, value=table.title).font = TITLE_FONT
    row += 1
    if table.subtitle:
        worksheet.cell(row=row, column=1, value=table.subtitle).font = SUBTITLE_FONT
        row += 1
    row += 1                                   # blank spacer

    header_row = row
    if table.group_spans:
        column = 1
        for label, span in table.group_spans:
            if label:
                cell = worksheet.cell(row=row, column=column, value=label)
                cell.font = GROUP_FONT
                cell.fill = GROUP_FILL
                cell.alignment = Alignment(horizontal="center")
                if span > 1:
                    worksheet.merge_cells(
                        start_row=row, start_column=column,
                        end_row=row, end_column=column + span - 1,
                    )
            column += span
        row += 1
        header_row = row

    for index, column_def in enumerate(table.columns, start=1):
        cell = worksheet.cell(row=row, column=index, value=column_def.header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        worksheet.column_dimensions[get_column_letter(index)].width = column_def.width
    row += 1

    first_data_row = row
    for data_row in table.rows:
        for index, cell in enumerate(data_row, start=1):
            target = worksheet.cell(row=row, column=index, value=_excel_value(cell))
            target.number_format = NUMBER_FORMATS.get(cell.style, "@")
            target.border = BORDER
            if cell.style == "code" and cell.value:
                target.alignment = Alignment(horizontal="center")
                target.fill = CODE_FILLS.get(str(cell.value), LEAVE_FILL)
            elif cell.style in {"time", "duration", "int", "number", "percent"}:
                target.alignment = Alignment(horizontal="center")
        row += 1

    if table.rows:
        worksheet.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(width)}{row - 1}"
        )
    worksheet.freeze_panes = worksheet.cell(row=first_data_row, column=4)

    if table.notes:
        row += 1
        for note in table.notes:
            worksheet.cell(row=row, column=1, value=note).font = NOTE_FONT
            row += 1
    return row + 2


def _safe_sheet_title(name: str, used: set[str]) -> str:
    cleaned = "".join(ch for ch in name if ch not in "[]:*?/\\")[:31] or "Sheet"
    candidate = cleaned
    suffix = 2
    while candidate in used:
        candidate = f"{cleaned[:28]}_{suffix}"
        suffix += 1
    used.add(candidate)
    return candidate


def write_workbook(
    tables: Sequence[ReportTable] | ReportTable,
    path: str | Path,
    *,
    sheet_names: Sequence[str] | None = None,
    combine: bool = False,
    combined_sheet_name: str = "Summary",
) -> Path:
    """Write one or more tables to ``path``.

    ``combine=True`` stacks every table on a single sheet (used for the daily
    exception summary, whose sections belong together).
    """
    if isinstance(tables, ReportTable):
        tables = [tables]
    tables = list(tables)
    if not tables:
        raise ValueError("nothing to write: no report tables supplied")

    workbook = Workbook()
    workbook.remove(workbook.active)
    used: set[str] = set()

    if combine:
        worksheet = workbook.create_sheet(_safe_sheet_title(combined_sheet_name, used))
        row = 1
        for table in tables:
            row = _write_table(worksheet, table, start_row=row)
    else:
        for index, table in enumerate(tables):
            name = (
                sheet_names[index]
                if sheet_names and index < len(sheet_names)
                else (table.meta.get("sheet_name") or table.key or f"Report{index + 1}")
            )
            worksheet = workbook.create_sheet(_safe_sheet_title(str(name), used))
            _write_table(worksheet, table)

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _csv_value(cell: Cell) -> str:
    if cell.value is None or cell.value == "":
        return ""
    if cell.style == "duration":
        return hhmm(cell.value) if isinstance(cell.value, timedelta) else str(cell.value)
    if cell.style == "time" and isinstance(cell.value, datetime):
        return cell.value.strftime("%Y-%m-%d %H:%M")
    if cell.style == "date" and isinstance(cell.value, date):
        return cell.value.isoformat()
    if cell.style == "percent" and isinstance(cell.value, (int, float)):
        return f"{cell.value * 100:.1f}%"
    return str(cell.value)


def write_csv(table: ReportTable, path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow([column.header for column in table.columns])
        for row in table.rows:
            writer.writerow([_csv_value(cell) for cell in row])
    return path


def write_csvs(tables: Iterable[ReportTable], directory: str | Path) -> list[Path]:
    directory = Path(directory)
    return [write_csv(table, directory / f"{table.key}.csv") for table in tables]
